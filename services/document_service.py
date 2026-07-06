import fitz  # PyMuPDF
from docx import Document
from openpyxl import load_workbook
from io import BytesIO


class DocumentService:

    def extract_text(self, uploaded_file):
        """
        Extract text based on the uploaded file type.
        """
        file_name = uploaded_file.name.lower()

        if file_name.endswith(".pdf"):
            return self.extract_pdf_text(uploaded_file)

        elif file_name.endswith(".docx"):
            return self.extract_docx_text(uploaded_file)

        elif file_name.endswith(".xlsx"):
            return self.extract_xlsx_text(uploaded_file)

        else:
            raise ValueError("Unsupported file type.")

    def extract_pdf_text(self, uploaded_file):
        pdf = fitz.open(
            stream=uploaded_file.read(),
            filetype="pdf"
        )

        text = ""

        for page in pdf:
            text += page.get_text()

        pdf.close()

        return text

    def extract_docx_text(self, uploaded_file):
        document = Document(BytesIO(uploaded_file.read()))

        text = []

        for paragraph in document.paragraphs:
            if paragraph.text.strip():
                text.append(paragraph.text)

        return "\n".join(text)

    def extract_xlsx_text(self, uploaded_file):
        workbook = load_workbook(
            filename=BytesIO(uploaded_file.read()),
            data_only=True
        )

        text = []

        for sheet in workbook.worksheets:
            text.append(f"\n=== Sheet: {sheet.title} ===\n")

            for row in sheet.iter_rows(values_only=True):
                values = [
                    str(cell) for cell in row
                    if cell is not None
                ]

                if values:
                    text.append(" | ".join(values))

        return "\n".join(text)